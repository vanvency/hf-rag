from pathlib import Path

import typer
from rich.console import Console

from src.core.config import get_settings
from src.core.logging import configure_logging
from src.retrieval.search import VectorStore
from src.services.processor import DocumentProcessor

app = typer.Typer(help="Document processing pipeline")
console = Console()


@app.command()
def run(directory: str = "data/upload"):
    """Process files from the upload directory and update the vector store."""
    settings = get_settings()
    logger = configure_logging()
    vector_store = VectorStore(settings.vector_store_path)
    processor = DocumentProcessor(settings=settings, vector_store=vector_store, logger=logger)

    processed = processor.process_directory(Path(directory))
    docs, chunks = vector_store.stats()
    console.print(
        f"[green]Processed {processed} new files. Store now tracks {docs} documents / {chunks} chunks.[/green]"
    )


@app.command()
def testcases(
    file_path: str = "tests/testcases.xlsx",
    api_url: str = "http://localhost:8000/api/query",
    top_k: int = 3,
    threshold: float = 0.0,
    delay: float = 0.2,
):
    """Run questions from an Excel file against the query API and print results."""
    import re
    import requests
    from openpyxl import load_workbook

    # Load workbook and first sheet
    wb = load_workbook(filename=file_path)
    sheet = wb.active

    # Detect the question column
    header_cells = [cell.value for cell in sheet[1]]
    question_col_idx = 1
    if header_cells:
        for idx, val in enumerate(header_cells, start=1):
            text = str(val or "").strip().lower()
            if re.search(r"question|问题|query", text):
                question_col_idx = idx
                break

    # Iterate rows and query API
    total = 0
    ok = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        question = row[question_col_idx - 1]
        if not question or not str(question).strip():
            continue
        total += 1
        payload = {"query": str(question).strip(), "top_k": top_k, "threshold": threshold}
        try:
            resp = requests.post(
                api_url,
                json=payload,
                timeout=60,
                proxies={"http": None, "https": None},
                headers={"Content-Type": "application/json"},
            )
            resp.raise_for_status()
            data = resp.json()
            results = data.get("results", [])
            if results:
                ok += 1
                best = results[0]
                console.print(
                    f"[blue]Q:[/blue] {payload['query']}\n[green]Top1 score:[/green] {best.get('score')}\n[cyan]Content:[/cyan] {best.get('content')[:200]}\n"
                )
            else:
                console.print(f"[yellow]Q:[/yellow] {payload['query']} -> [red]No results[/red]")
        except Exception as e:
            console.print(f"[red]Error querying[/red] '{payload['query']}': {e}")
        import time
        time.sleep(delay)

    console.print(f"[bold]Ran {total} questions, {ok} returned results[/bold]")


if __name__ == "__main__":
    app()

